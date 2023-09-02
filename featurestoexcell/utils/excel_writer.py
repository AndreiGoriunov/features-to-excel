from os import makedirs, path

from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from featurestoexcell.utils import get_datetime


class ExcelWriter:
    """
    A class to manage the writing and styling of data into Excel workbooks.
    
    Attributes:
    - output_dir: The directory where the workbook will be saved.
    - max_column_width: The maximum allowed width for any column.
    - wb: The active workbook.
    """
    def __init__(self, output_dir: str, max_column_width: int = 70) -> None:
        """
        Initializes the ExcelWriter with the given output directory and maximum column width.
        
        Args:
        - output_dir: Directory to save the resulting workbook.
        - max_column_width: The maximum width a column can be set to.
        """
        self.output_dir = path.normpath(output_dir)
        self.max_column_width = max_column_width
        self.wb = Workbook()
        self.wb.active.title = "Summary"  # type: ignore

    def _apply_styling(self, ws: Worksheet):
        """
        Adjust the width of columns and height of rows in a worksheet based on their content.
        Wraps text in all cells, and ensures that columns do not exceed a set maximum width.
        
        Args:
        - ws: The worksheet to be styled.
        """
        max_width = self.max_column_width
        for col in ws.columns:
            max_length = 0
            column = [cell for cell in col] # type: ignore
            for cell in column:
                # Apply text wrap to all cells
                cell.alignment = Alignment(wrap_text=True)
                
                # Calculate the number of lines in a cell to adjust row height
                number_of_lines = cell.value.count('\n') + 1
                ws.row_dimensions[cell.row].height = number_of_lines * 15  # assuming a height of 15 units per line

                # Determine the maximum cell width for this column
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass

            # Adjust the column width based on the maximum cell width, ensuring it doesn't exceed the max width
            adjusted_width = (max_length + 2)  # a little extra space
            if adjusted_width > max_width:  # check against the max width
                adjusted_width = max_width
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    def save_workbook(self, base_name: str):
        """
        Save the workbook after applying styling to all its sheets.
        
        Args:
        - base_name: The base filename for the saved workbook.
        """
        # Adjust column widths and row heights for each worksheet
        for ws in self.wb.worksheets:
            self._apply_styling(ws)

        # Create output directory if it doesn't exist
        makedirs(self.output_dir, exist_ok=True)

        # Determine the complete file path and save the workbook
        save_filepath: str = path.normpath(
            path.join(self.output_dir, f"{base_name}_{get_datetime()}.xlsx")
        )
        self.wb.save(save_filepath)

    def write_to_worksheet(self, name: str, headers: list[str], values: list[str]):
        """
        Writes headers and values to the specified worksheet. If the worksheet does not exist, it will be created.

        Args:
        - name: Name of the worksheet to write to.
        - headers: Column headers.
        - values: Data values.
        """
        # Check if the worksheet by the given name already exists
        sheet_names = [ws.title for ws in self.wb.worksheets]
        if not name in sheet_names:
            # Create a new worksheet with the given name and append headers to it
            ws: Worksheet = self.wb.create_sheet(name)
            ws.append(headers)

        # If it exists, get the worksheet
        ws = self.wb[name]
        # Append the given values to the worksheet
        ws.append(values)